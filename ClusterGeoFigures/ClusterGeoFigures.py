import os
import numpy as np
import matplotlib.colors as mcolors

from ase.io import read, write
from ase.visualize import view

#from ase.data import atomic_numbers, chemical_symbols
#from asap3.analysis.rdf import RadialDistributionFunction

from collections import Counter

import matplotlib.pyplot as plt
from matplotlib.pyplot import figure

from ClusterGeoFigures.No_Of_Neighbours import No_Of_Neighbours

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

class ClusterGeoFigures_Program:
	def __init__(self, r_cut, elements=['Cu','Pd'],focus_plot_with_respect_to_element='Cu',path_to_xyz_files='.',add_legend=False,bulk_colour='#FFC0CB',face_colour='#FF0000',vertex_colour='#90EE90',edge_colour='#ADD8E6',none_colour='#FFFFFF'):
		self.path_to_xyz_files = os.path.abspath(path_to_xyz_files)
		self.r_cut = r_cut
		self.elements = elements
		self.focus_plot_with_respect_to_element = focus_plot_with_respect_to_element
		self.add_legend = add_legend

		def rgb_to_hex(rgb):
			r, g, b = rgb
			r, g, b = int(r*255.0), int(g*255.0), int(b*255.0)
			return '#%02x%02x%02x' % (r, g, b)

		self.bulk_colour = (bulk_colour if isinstance(bulk_colour,str) else rgb_to_hex(bulk_colour)).upper() 
		self.face_colour = (face_colour if isinstance(face_colour,str) else rgb_to_hex(face_colour)).upper() 
		self.vertex_colour = (vertex_colour if isinstance(vertex_colour,str) else rgb_to_hex(vertex_colour)).upper() 
		self.edge_colour = (edge_colour if isinstance(edge_colour,str) else rgb_to_hex(edge_colour)).upper() 
		self.none_colour = (none_colour if isinstance(none_colour,str) else rgb_to_hex(none_colour)).upper() 
		self.colours = {'bulk': self.bulk_colour, 'face': self.face_colour, 'edge': self.edge_colour, 'vertex': self.vertex_colour, 'None': self.none_colour}

		self.to_save = []
		self.to_save += [str(self.focus_plot_with_respect_to_element)+'_bulk'  ,'bulk'  ,str(self.focus_plot_with_respect_to_element)+'_percent_bulk']
		self.to_save += [str(self.focus_plot_with_respect_to_element)+'_face','face'  ,str(self.focus_plot_with_respect_to_element)+'_percent_face']
		self.to_save += [str(self.focus_plot_with_respect_to_element)+'_edge'  ,'edge'  ,str(self.focus_plot_with_respect_to_element)+'_percent_edge']
		self.to_save += [str(self.focus_plot_with_respect_to_element)+'_vertex','vertex',str(self.focus_plot_with_respect_to_element)+'_percent_vertex']

		self.created_figures_folder = 'ClusterGeoFigures_figures_folder'
		if not os.path.exists(self.created_figures_folder):
			os.makedirs(self.created_figures_folder)
		self.run()

	def run(self):
		print('--------------------------------------')
		print('Getting data from XYZ files')
		clusters_data = self.get_cluster_data()
		print('Finished getting data from XYZ files')
		print('--------------------------------------')
		print('Analysing data')
		cluster_information = self.analyse_cluster_data(clusters_data)
		print('Finished analysing data')
		print('--------------------------------------')
		print('Making plots')
		cluster_plot_data = self.make_plots(cluster_information)
		print('Finished making plots')
		print('--------------------------------------')
		print('Making excel spreadsheet')
		self.record_to_excel(cluster_plot_data)
		print('Finished excel spreadsheet')
		print('--------------------------------------')

	def get_cluster_data(self):
		clusters_data = {}
		for file in os.listdir(self.path_to_xyz_files):
			if not file.endswith('.xyz'):
				continue
			cluster = read(self.path_to_xyz_files+'/'+file)
			clusters_data.setdefault(len(cluster),[]).append((cluster,file))
		self.greatest_no_of_atoms = max(clusters_data.keys())
		clusters_data_temp = clusters_data[self.greatest_no_of_atoms]
		clusters_data = {}
		for cluster, file in clusters_data_temp:
			symbols = cluster.get_chemical_symbols()
			no_of_symbols = Counter(symbols)
			name = ''.join([element+str(no_of_symbols[element]) for element in self.elements])+'.xyz'
			name = name.lower()
			if not name.lower() == file.lower():
				continue
			clusters_data[tuple([no_of_symbols[element] for element in self.elements])] = cluster
		return clusters_data

	def analyse_cluster_data(self,clusters_data):
		cluster_information = self.analyse_NN_1(clusters_data)
		return cluster_information

	def analyse_NN_1(self, clusters_data):
		cluster_information = {}
		for key in sorted(clusters_data.keys()):
			cluster = clusters_data[key]
			nl = No_Of_Neighbours([self.r_cut/2.0]*len(cluster))
			nl.update(cluster)
			all_number_of_neighbours = {}
			element_number_of_neighbours = {}
			for index in range(len(cluster)):
				indices, offsets = nl.get_neighbors(index)
				number_of_neighbours = len(indices)
				all_number_of_neighbours.setdefault(number_of_neighbours,[]).append(index)
				if cluster[index].symbol == self.focus_plot_with_respect_to_element:
					element_number_of_neighbours.setdefault(number_of_neighbours,[]).append(index)
			cluster_information[key] = (element_number_of_neighbours, all_number_of_neighbours, cluster)
		return cluster_information

	#-----------------------------------------------------------------------------------------------------------------------
	#-----------------------------------------------------------------------------------------------------------------------
	#-----------------------------------------------------------------------------------------------------------------------

	def make_plots(self, cluster_information_overall):
		cluster_information = [(key, element_number_of_neighbours, all_number_of_neighbours, cluster) for key, (element_number_of_neighbours, all_number_of_neighbours, cluster) in cluster_information_overall.items()]
		cluster_information.sort()
		cluster_plot_data = {}
		no_of_copper_atoms = []

		bulk_data   = [] # 12+ neighbours
		face_data   = [] # 9-11 neighbours
		edge_data   = [] # 7-8 neighbours
		vertex_data = [] # <= 6 nieghbours

		max_bulk_data   = [] # 12+ neighbours
		max_face_data   = [] # 9-11 neighbours
		max_edge_data   = [] # 7-8 neighbours
		max_vertex_data = [] # <= 6 nieghbours
		
		norm_bulk_data   = [] # 12+ neighbours
		norm_face_data   = [] # 9-11 neighbours
		norm_edge_data   = [] # 7-8 neighbours
		norm_vertex_data = [] # <= 6 nieghbours
		for key, element_number_of_neighbours, all_number_of_neighbours, cluster in cluster_information:
			no_of_element_atoms_in_cluster = key[self.elements.index(self.focus_plot_with_respect_to_element)]
			no_of_copper_atoms.append(no_of_element_atoms_in_cluster)
			# ---------------------------------------------------------------------
			element_bulk = element_face = element_edge = element_vertex = 0
			for number_of_neighbours, indices in element_number_of_neighbours.items():
				if number_of_neighbours >= 12:
					element_bulk += len(indices)
				elif 9 <= number_of_neighbours <= 11:
					element_face += len(indices)
				elif 7 <= number_of_neighbours <= 8:
					element_edge += len(indices)
				elif number_of_neighbours <= 6:
					element_vertex += len(indices)
				else:
					exit('Huh?')
			# ---------------------------------------------------------------------
			bulk = face = edge = vertex = 0
			for number_of_neighbours, indices in all_number_of_neighbours.items():
				if number_of_neighbours >= 12:
					bulk += len(indices)
				elif 9 <= number_of_neighbours <= 11:
					face += len(indices)
				elif 7 <= number_of_neighbours <= 8:
					edge += len(indices)
				elif number_of_neighbours <= 6:
					vertex += len(indices)
				else:
					exit('Huh?')
			# ---------------------------------------------------------------------
			element_percent_bulk   = (element_bulk/bulk)*100.0     if not bulk == 0 else -50.0
			element_percent_face   = (element_face/face)*100.0     if not face == 0 else -50.0
			element_percent_edge   = (element_edge/edge)*100.0     if not edge == 0 else -50.0
			element_percent_vertex = (element_vertex/vertex)*100.0 if not vertex == 0 else -50.0
			# ---------------------------------------------------------------------
			print('no_of_'+str(self.focus_plot_with_respect_to_element)+'_atoms_in_cluster: '+str(no_of_element_atoms_in_cluster))
			print(str(self.focus_plot_with_respect_to_element)+'_bulk: '+str(element_bulk)+'; bulk: '+str(bulk)+': Percent: '+str(element_percent_bulk))
			print(str(self.focus_plot_with_respect_to_element)+'_face: '+str(element_face)+'; face: '+str(face)+': Percent: '+str(element_percent_face))
			print(str(self.focus_plot_with_respect_to_element)+'_edge: '+str(element_edge)+'; edge: '+str(edge)+': Percent: '+str(element_percent_edge))
			print(str(self.focus_plot_with_respect_to_element)+'_vertex: '+str(element_vertex)+'; vertex: '+str(vertex)+': Percent: '+str(element_percent_vertex))
			print('--------------------')
			# ---------------------------------------------------------------------
			bulk_data.append(element_bulk if bulk > 0 else -50)
			face_data.append(element_face if face > 0 else -50)
			edge_data.append(element_edge if edge > 0 else -50)
			vertex_data.append(element_vertex if vertex > 0 else -50)

			max_bulk_data.append(bulk if bulk > 0 else -50)
			max_face_data.append(face if face > 0 else -50)
			max_edge_data.append(edge if edge > 0 else -50)
			max_vertex_data.append(vertex if vertex > 0 else -50)
			
			norm_bulk_data.append(element_percent_bulk)
			norm_face_data.append(element_percent_face)
			norm_edge_data.append(element_percent_edge)
			norm_vertex_data.append(element_percent_vertex)
			# ---------------------------------------------------------------------
			cluster_plot_data[key] = {}
			for data_name in self.to_save:
				if self.focus_plot_with_respect_to_element in data_name:
					getting_data_name = data_name.replace(self.focus_plot_with_respect_to_element,'element')
				else:
					getting_data_name = data_name
				cluster_plot_data[key][data_name] = eval(getting_data_name)
			# ---------------------------------------------------------------------
		# ---------------------------------------------------------------------
		# make figures
		norm_figure_data = [norm_vertex_data, norm_edge_data, norm_face_data, norm_bulk_data]
		full_figure_data = [max_vertex_data, max_edge_data, max_face_data, max_bulk_data, vertex_data, edge_data, face_data, bulk_data]
		for show_figure, figure_data in (('normalised_figures',norm_figure_data), ('no_of_elements_given_as_no_of_atoms',full_figure_data), ('no_of_elements_given_as_percent_figures', full_figure_data)):
			self.make_single_figure(show_figure=show_figure,no_of_copper_atoms=no_of_copper_atoms,figure_data=figure_data)
			for figure_format in ('square','long','wide'):
				self.make_separated_figure(show_figure=show_figure,figure_format=figure_format,no_of_copper_atoms=no_of_copper_atoms,figure_data=figure_data)
		# ---------------------------------------------------------------------
		return cluster_plot_data

	def make_single_figure(self,show_figure,no_of_copper_atoms,figure_data):
		# -------------------------------------------------------------------------
		# -------------------------------------------------------------------------
		figure(figsize=(8, 3), dpi=80)
		if   show_figure == 'normalised_figures':
			# normalised percents of elements in positions in cluster

			norm_vertex_data, norm_edge_data, norm_face_data, norm_bulk_data = figure_data

			plt.plot([0, self.greatest_no_of_atoms], [0, 100], c='k', ls='-', lw=2.0)

			plt.scatter(no_of_copper_atoms,norm_vertex_data,c=self.colours['vertex'],label='vertex')
			plt.scatter(no_of_copper_atoms,norm_edge_data,c=self.colours['edge'],label='edge')
			plt.scatter(no_of_copper_atoms,norm_face_data,c=self.colours['face'],label='face')
			plt.scatter(no_of_copper_atoms,norm_bulk_data,c=self.colours['bulk'],label='bulk')

			
			plt.ylabel('Normalised % '+str(self.focus_plot_with_respect_to_element)+' composition')
			plt.ylim((-0.5,100.5))
			plt.xlim((0-1,self.greatest_no_of_atoms+1))

			name_of_file = 'normalised_percent_'

		elif show_figure == 'no_of_elements_given_as_no_of_atoms':
			# no of elements in positions in cluster, with maximum number of those positions.
			# given as the number of atoms

			max_vertex_data, max_edge_data, max_face_data, max_bulk_data, vertex_data, edge_data, face_data, bulk_data = figure_data

			plt.scatter(no_of_copper_atoms,max_vertex_data,edgecolors=self.colours['vertex'], c='white', linewidths=2,label='max vertex')
			plt.scatter(no_of_copper_atoms,max_edge_data,edgecolors=self.colours['edge'], c='white', linewidths=2,label='max edge')
			plt.scatter(no_of_copper_atoms,max_face_data,edgecolors=self.colours['face'], c='white', linewidths=2,label='max face')
			plt.scatter(no_of_copper_atoms,max_bulk_data,edgecolors=self.colours['bulk'], c='white', linewidths=2,label='max bulk')

			plt.scatter(no_of_copper_atoms,vertex_data,c=self.colours['vertex'],label='vertex')
			plt.scatter(no_of_copper_atoms,edge_data,c=self.colours['edge'],label='edge')
			plt.scatter(no_of_copper_atoms,face_data,c=self.colours['face'],label='face')
			plt.scatter(no_of_copper_atoms,bulk_data,c=self.colours['bulk'],label='bulk')

			plt.ylabel('# '+str(self.focus_plot_with_respect_to_element)+' atoms composition')
			plt.ylim((-0.5,self.greatest_no_of_atoms+.5))
			plt.xlim((0-1,self.greatest_no_of_atoms+1))

			name_of_file = 'no_of_'

		elif show_figure == 'no_of_elements_given_as_percent_figures':
			# no of elements in positions in cluster, with maximum number of those positions.
			# given as percentage of number of atoms in cluster

			max_vertex_data, max_edge_data, max_face_data, max_bulk_data, vertex_data, edge_data, face_data, bulk_data = figure_data

			plt.scatter(no_of_copper_atoms,np.array(max_vertex_data)*(100.0/float(self.greatest_no_of_atoms)),edgecolors=self.colours['vertex'], c='white', linewidths=2,label='max vertex')
			plt.scatter(no_of_copper_atoms,np.array(max_edge_data)*(100.0/float(self.greatest_no_of_atoms)),edgecolors=self.colours['edge'], c='white', linewidths=2,label='max edge')
			plt.scatter(no_of_copper_atoms,np.array(max_face_data)*(100.0/float(self.greatest_no_of_atoms)),edgecolors=self.colours['face'], c='white', linewidths=2,label='max face')
			plt.scatter(no_of_copper_atoms,np.array(max_bulk_data)*(100.0/float(self.greatest_no_of_atoms)),edgecolors=self.colours['bulk'], c='white', linewidths=2,label='max bulk')

			plt.scatter(no_of_copper_atoms,np.array(vertex_data)*(100.0/float(self.greatest_no_of_atoms)),c=self.colours['vertex'],label='vertex')
			plt.scatter(no_of_copper_atoms,np.array(edge_data)*(100.0/float(self.greatest_no_of_atoms)),c=self.colours['edge'],label='edge')
			plt.scatter(no_of_copper_atoms,np.array(face_data)*(100.0/float(self.greatest_no_of_atoms)),c=self.colours['face'],label='face')
			plt.scatter(no_of_copper_atoms,np.array(bulk_data)*(100.0/float(self.greatest_no_of_atoms)),c=self.colours['bulk'],label='bulk')

			plt.ylabel('% '+str(self.focus_plot_with_respect_to_element)+' composition')
			plt.ylim((-0.5,100.5))
			plt.xlim((0-1,self.greatest_no_of_atoms+1))

			name_of_file = 'percent_'

		plt.xlabel('$\mathregular{N_{'+str(self.focus_plot_with_respect_to_element)+'}}$')
		plt.tight_layout()
		if self.add_legend:
			plt.legend()

		name_of_file += str(self.focus_plot_with_respect_to_element)+'_comp_vs_no_'+str(self.focus_plot_with_respect_to_element)+'_atoms'
		plt.savefig(self.created_figures_folder+'/'+name_of_file+'.png')
		plt.savefig(self.created_figures_folder+'/'+name_of_file+'.svg')

	def make_separated_figure(self,show_figure,figure_format,no_of_copper_atoms,figure_data):

		if figure_format == 'square':
			fig = figure(figsize=(8, 6), dpi=80)
			ax = fig.add_subplot(111)    # The big subplot
			ax1 = fig.add_subplot(221)
			ax2 = fig.add_subplot(222)
			ax3 = fig.add_subplot(223)
			ax4 = fig.add_subplot(224)
		elif figure_format == 'long':
			fig = figure(figsize=(4, 11), dpi=80)
			ax = fig.add_subplot(111)    # The big subplot
			ax1 = fig.add_subplot(411)
			ax2 = fig.add_subplot(412)
			ax3 = fig.add_subplot(413)
			ax4 = fig.add_subplot(414)
		elif figure_format == 'wide':
			fig = figure(figsize=(15, 4), dpi=80)
			ax = fig.add_subplot(111)    # The big subplot
			ax1 = fig.add_subplot(141)
			ax2 = fig.add_subplot(142)
			ax3 = fig.add_subplot(143)
			ax4 = fig.add_subplot(144)
		else:
			print('Error in make_separated_figure method')
			exit('exitting')

		# Turn off axis lines and ticks of the big subplot
		ax.spines['top'].set_color('none')
		ax.spines['bottom'].set_color('none')
		ax.spines['left'].set_color('none')
		ax.spines['right'].set_color('none')
		ax.tick_params(labelcolor='w', top=False, bottom=False, left=False, right=False)

		# ---------------------------------

		if show_figure == 'normalised_figures':
			# normalised percents of elements in positions in cluster

			norm_vertex_data, norm_edge_data, norm_face_data, norm_bulk_data = figure_data

			for axis in [ax1, ax2, ax3, ax4]:
				axis.plot([0, self.greatest_no_of_atoms], [0, 100], c='k', ls='-', lw=2.0, zorder=-10)
				axis.set_ylim((-0.5,100.5))
				axis.set_xlim((0-1,self.greatest_no_of_atoms+1))

			ax4.scatter(no_of_copper_atoms,norm_vertex_data,c=self.colours['vertex'],label='vertex')
			ax3.scatter(no_of_copper_atoms,norm_edge_data,c=self.colours['edge'],label='edge')
			ax2.scatter(no_of_copper_atoms,norm_face_data,c=self.colours['face'],label='face')
			ax1.scatter(no_of_copper_atoms,norm_bulk_data,c=self.colours['bulk'],label='bulk')

			ax.set_ylabel('Normalised % '+str(self.focus_plot_with_respect_to_element)+' composition')

			name_of_file = 'normalised_percent_'

		elif show_figure == 'no_of_elements_given_as_no_of_atoms':
			# no of elements in positions in cluster, with maximum number of those positions.
			# given as the number of atoms

			max_vertex_data, max_edge_data, max_face_data, max_bulk_data, vertex_data, edge_data, face_data, bulk_data = figure_data

			for axis in [ax1, ax2, ax3, ax4]:
				axis.set_ylim((0-1,self.greatest_no_of_atoms+1))
				axis.set_xlim((0-1,self.greatest_no_of_atoms+1))

			ax4.scatter(no_of_copper_atoms,max_vertex_data,edgecolors=self.colours['vertex'], c='white', linewidths=2,label='max vertex')
			ax3.scatter(no_of_copper_atoms,max_edge_data,edgecolors=self.colours['edge'], c='white', linewidths=2,label='max edge')
			ax2.scatter(no_of_copper_atoms,max_face_data,edgecolors=self.colours['face'], c='white', linewidths=2,label='max face')
			ax1.scatter(no_of_copper_atoms,max_bulk_data,edgecolors=self.colours['bulk'], c='white', linewidths=2,label='max bulk')

			ax4.scatter(no_of_copper_atoms,vertex_data,c=self.colours['vertex'],label='vertex')
			ax3.scatter(no_of_copper_atoms,edge_data,c=self.colours['edge'],label='edge')
			ax2.scatter(no_of_copper_atoms,face_data,c=self.colours['face'],label='face')
			ax1.scatter(no_of_copper_atoms,bulk_data,c=self.colours['bulk'],label='bulk')

			ax.set_ylabel('# '+str(self.focus_plot_with_respect_to_element)+' atoms composition')

			name_of_file = 'no_of_'

		elif show_figure == 'no_of_elements_given_as_percent_figures':
			# no of elements in positions in cluster, with maximum number of those positions.
			# given as percentage of number of atoms in cluster

			for axis in [ax1, ax2, ax3, ax4]:
				axis.set_ylim((-0.5,100.5))
				axis.set_xlim((0-1,self.greatest_no_of_atoms+1))

			max_vertex_data, max_edge_data, max_face_data, max_bulk_data, vertex_data, edge_data, face_data, bulk_data = figure_data

			ax4.scatter(no_of_copper_atoms,np.array(max_vertex_data)*(100.0/float(self.greatest_no_of_atoms)),edgecolors=self.colours['vertex'], c='white', linewidths=2,label='max vertex')
			ax3.scatter(no_of_copper_atoms,np.array(max_edge_data)*(100.0/float(self.greatest_no_of_atoms)),edgecolors=self.colours['edge'], c='white', linewidths=2,label='max edge')
			ax2.scatter(no_of_copper_atoms,np.array(max_face_data)*(100.0/float(self.greatest_no_of_atoms)),edgecolors=self.colours['face'], c='white', linewidths=2,label='max face')
			ax1.scatter(no_of_copper_atoms,np.array(max_bulk_data)*(100.0/float(self.greatest_no_of_atoms)),edgecolors=self.colours['bulk'], c='white', linewidths=2,label='max bulk')

			ax4.scatter(no_of_copper_atoms,np.array(vertex_data)*(100.0/float(self.greatest_no_of_atoms)),c=self.colours['vertex'],label='vertex')
			ax3.scatter(no_of_copper_atoms,np.array(edge_data)*(100.0/float(self.greatest_no_of_atoms)),c=self.colours['edge'],label='edge')
			ax2.scatter(no_of_copper_atoms,np.array(face_data)*(100.0/float(self.greatest_no_of_atoms)),c=self.colours['face'],label='face')
			ax1.scatter(no_of_copper_atoms,np.array(bulk_data)*(100.0/float(self.greatest_no_of_atoms)),c=self.colours['bulk'],label='bulk')
			
			ax.set_ylabel('% '+str(self.focus_plot_with_respect_to_element)+' composition')

			name_of_file = 'percent_'

		# ---------------------------------
		ax.set_xlabel('$\mathregular{N_{'+str(self.focus_plot_with_respect_to_element)+'}}$')
		plt.tight_layout()
		if self.add_legend:
			for axis in [ax1, ax2, ax3, ax4]:
				axis.legend()

		name_of_file += str(self.focus_plot_with_respect_to_element)+'_comp_vs_no_'+str(self.focus_plot_with_respect_to_element)+'_atoms'
		name_of_file += '_subplots'
		name_of_file += '_'+str(figure_format)

		plt.savefig(self.created_figures_folder+'/'+name_of_file+'.png')
		plt.savefig(self.created_figures_folder+'/'+name_of_file+'.svg')

	#-----------------------------------------------------------------------------------------------------------------------
	#-----------------------------------------------------------------------------------------------------------------------
	#-----------------------------------------------------------------------------------------------------------------------

	def record_to_excel(self, cluster_plot_data):
		cluster_plot_data = [(key, data) for key, data in cluster_plot_data.items()]
		cluster_plot_data.sort()

		workbook = Workbook()
		worksheet = workbook.active

		def hex_to_rgb(hex_key):
			rgb = []
			for i in (0,2,4):
				decimal = int(hex_key[i:i+2],16)
				rgb.append(decimal)
			return tuple(rgb)

		def give_black_or_white_writing(hex_key):
			rgb = hex_to_rgb(hex_key)
			record = []
			for c in rgb:
				c = c/255.0
				if c <= 0.03928:
					c = c/12.92
				else:
					c = ((c+0.055)/1.055) ** 2.4
				record.append(c)
			r, g, b = record
			L = 0.2126*r + 0.7152*g + 0.0722*b

			if L > ((1.05*0.05)**0.5 - 0.05):
				return '000000'
			else:
				return 'FFFFFF'

		def get_colour_name(name):
			for colour_name in self.colours.keys():
				if colour_name in name:
					return colour_name
			return 'None'

		naming = [str(tuple(self.elements))]+self.to_save
		for index in range(len(naming)):
			name = naming[index]
			worksheet.cell(column=index+1, row=1, value=str(name))
			colour_of_background = self.colours[get_colour_name(name)].replace('#','')
			worksheet.cell(column=index+1, row=1).font = Font(color=give_black_or_white_writing(colour_of_background))
			worksheet.cell(column=index+1, row=1).fill = PatternFill("solid", fgColor=colour_of_background)

		for index in range(len(cluster_plot_data)):
			key, data = cluster_plot_data[index]
			worksheet.cell(column=1, row=index+2, value=str(key))
			for index2 in range(len(self.to_save)):
				name = self.to_save[index2]
				if 'percent' in name:
					number = round(data[name],1)
					if number < 0:
						number = ''
				else:
					number = data[name]
				worksheet.cell(column=index2+2, row=index+2, value=str(number))
				colour_of_background = self.colours[get_colour_name(name)].replace('#','')
				worksheet.cell(column=index2+2, row=index+2).font = Font(color=give_black_or_white_writing(colour_of_background))
				worksheet.cell(column=index2+2, row=index+2).fill = PatternFill("solid", fgColor=colour_of_background)
		# Save the file
		workbook.save("ClusterGeo_Data_"+str(self.focus_plot_with_respect_to_element)+".xlsx")
